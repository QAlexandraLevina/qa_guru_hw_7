import csv
import os
import zipfile
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
from io import TextIOWrapper
from constants import FILES_DIR


"""Список файлов"""
file_for_zip = [
    "filepdf.pdf",
    "filexlsx.xlsx",
    "filecsv.csv",
]

"""Открываем архив и сохраняем в него файлы"""


def test_open_archive_save_files():
    with zipfile.ZipFile(f"{FILES_DIR}/archive.zip", "w") as zf:
        for file in file_for_zip:
            file_path = os.path.join(FILES_DIR, file)
            zf.write(file_path, arcname=file)
    assert os.path.exists(f"{FILES_DIR}/archive.zip")
    assert zipfile.is_zipfile(f"{FILES_DIR}/archive.zip")


"""Проверка наличия файлов в архиве"""


def test_archive_files():
    with ZipFile(f"{FILES_DIR}/archive.zip", "r") as zf:
        archive_files = zf.namelist()
        for file in file_for_zip:
            assert file in archive_files, f"Файла {file} нет в архиве!"


"""Проверка и чтение PDF файла из архива"""


def test_pdf_read_file():
    with ZipFile(f"{FILES_DIR}/archive.zip", "r") as zf:
        with zf.open("filepdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            numb_of_page = len(reader.pages)
            print(f"Количество страниц в PDF файле: {numb_of_page}")
            page = reader.pages[0]
            text_in_pdf = page.extract_text()
            print(f"Текст из файла PDF: {text_in_pdf}")
            assert "PDF" in text_in_pdf, "Такого слова нет в файле!"


"""Проверка и чтение XLSX файла из архива"""


def test_xlsx_read_file():
    with ZipFile(f"{FILES_DIR}/archive.zip", "r") as zf:
        with zf.open("filexlsx.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            sheet_cell_value = sheet.cell(row=7, column=2).value
            print(f"Текст из ячейки: {sheet_cell_value}")
            assert "Chasse" in sheet_cell_value, "Такого значения нет в ячейке!"


"""Проверка и чтение CSV файла из архива"""


def test_csv_read_file():
    with ZipFile(f"{FILES_DIR}/archive.zip", "r") as zf:
        with zf.open("filecsv.csv") as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, "utf-8-sig")))
            third_row = csvreader[2]
            print(f"Текст из строки: {third_row}")
            assert third_row[0] == "Josefina Hessel"
            assert third_row[2] == "penelope.gerlach@hotmail.com"
